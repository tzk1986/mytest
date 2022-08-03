import os

import openpyxl as openpyxl


class ExcelUtil:
    # 获取项目路径
    def get_object_path(self):
        print(os.path.abspath(os.path.dirname(__file__)).split("common")[0])

    def read_execl(self):
        # openpyxl,xlrd  python中读取execl的第三方服务
        # 加载execl工作簿
        wb = openpyxl.load_workbook(self.get_object_path() + "execl存放路径")
        # 获得sheet对象
        sheet = wb['sheet名称']
        # 获得execl的行数和列数
        print(sheet.max_row, sheet.max_column)
        # 循环
        # for rows in range(1, sheet.max_row + 1):
        #     for cols in range(1, sheet.max_column + 1):
        #         print(rows, cols)  # 显示可读取的行和列数量
        #         print(sheet.cell(rows, cols).value)  # 显示行和列的数据
        # 把以上数据组成需要的形式
        all_list = []
        for rows in range(1, sheet.max_row + 1):
            temp_list = []
            for cols in range(1, sheet.max_column + 1):
                temp_list.append(sheet.cell(rows, cols).value)
            all_list.append(temp_list)
        return all_list

